import asyncio
import json
import shutil
import tempfile
import zipfile
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from nicegui import ui
from openpyxl import load_workbook

from apps.gui.components.widgets.metadataSettingsUi import MetadataSettingsUI
from apps.gui.helpers.io import spooled_file_to_temp_file
from imxInsights.file.singleFileImx.imxSituationEnum import ImxSituationEnum
from src.imxTools.insights.diff_and_population import write_diff_output_files
from src.imxTools.insights.measure_analyse import generate_measure_excel
from src.imxTools.revision.process_revision import process_imx_revisions
from src.imxTools.utils.helpers import load_imxinsights_container_or_file, create_timestamp


@dataclass
class MeasureCorrectionState:
    imx_file_path: Path | None = None
    loaded_imx_data: Any = None
    measure_excel_file: Path | None = None
    revision_log_zip: Path | None = None
    processed_imx: Path | None = None
    gr_json_file_path: Path | None = None
    revisions_excel_upload_widget: Path | None = None
    time_stamp: str | None = None


class MeasureCorrectionTool:
    UPLOAD_STEP = "Upload Required and Optional Files"
    REVIEW_STEP = "Review and Flag Revisions"
    CHECK_RESULT_STEP = "Check Revision Results"
    DEFAULT_SITUATION = ImxSituationEnum.NewSituation

    def __init__(self, container):
        self.state = MeasureCorrectionState()
        self._temp_dirs: list[str] = []
        with container:
            self._setup_widgets()

    def __del__(self):
        self._cleanup()

    def _setup_widgets(self):
        with ui.stepper().props("vertical").classes("w-full") as self.stepper:
            self._build_upload_step()
            self._build_review_step()
            self._build_check_result_step()

    def _build_upload_step(self):
        with ui.step(self.UPLOAD_STEP).classes("font-bold"):
            ui.label("Required:")
            self._build_imx_upload_card()

            ui.separator()
            ui.label("Optional").classes("italic")
            self._build_json_upload_card()

            with ui.stepper_navigation():
                self.analyze_measures_button = ui.button("Analyse Measures", on_click=self.run_measure_check)
                self.analyze_measures_button.disable()
                self.upload_spinner = ui.spinner(size="lg")
                self.upload_spinner.visible = False

    def _build_imx_upload_card(self):
        with ui.card().classes("w-full"):
            ui.label("The IMX file you want to analyze and correct.")
            self.imx_upload_widget = self._build_upload_widget(
                label="Upload IMX file", on_upload=self._on_imx_upload, allowed_extensions=[".xml", ".zip"]
            )
            ui.label('If a single IMX project file is uploaded, only the "NewSituation" will be corrected.')\
                .classes("text-sm italic text-gray-500")
            self.threshold_input_field = ui.number(label="Threshold (meters)", value=0.015).classes("w-full")

    def _build_json_upload_card(self):
        with ui.card().classes("w-full"):
            ui.label("The Naiade JSON file to exclude 'context area' objects.")
            ui.label("When uploaded, only 'work area', 'user area', and 'new' objects will be flagged.")\
                .classes("text-sm italic text-gray-500")
            self.gr_json_upload_widget = self._build_upload_widget(
                label="Upload the JSON file (GR_xxxxx.json)",
                on_upload=self._on_json_upload,
                allowed_extensions=[".json"],
            )
            self.exclude_context_checkbox = ui.checkbox(
                "Flag all 'work area', 'user area', and 'new' objects (excluding 'context area')."
            )
            self.exclude_context_checkbox.disable()

    def _build_review_step(self):
        with ui.step(self.REVIEW_STEP).classes("font-bold"):
            ui.label("The uploaded IMX file has been analyzed.").classes("text-xl font-semibold text-primary")
            ui.label('You can now download the analyse (Excel) and review detected discrepancies in the "revisions" sheet.')\
                .classes("text-sm")
            ui.label('If a correction is needed, set the value in the "processing required" column to `True`.')\
                .classes("text-sm")

            ui.button("Download Revisions File", icon="download", on_click=self._on_download_revisions)\
                .props("outline")

            ui.space()
            ui.label("After reviewing and correcting the revisions, upload the file to proceed.")

            self.revisions_excel_upload_widget = self._build_upload_widget(
                label="Upload flagged revisions Excel file",
                on_upload=self._on_revision_upload,
                allowed_extensions=[".xlsx"],
            )

            self.metadata_ui = MetadataSettingsUI()

            with ui.stepper_navigation():
                self.process_revisions_button = ui.button("Process Revisions", on_click=self.process_revisions)
                self.process_revisions_back = ui.button("Back", on_click=self.stepper.previous).props("flat")
                self.review_spinner = ui.spinner(size="lg")
                self.review_spinner.visible = False

    def _build_check_result_step(self):
        with ui.step(self.CHECK_RESULT_STEP).classes("font-bold"):
            ui.label("The IMX file has been updated.").classes("text-xl font-semibold text-primary")
            ui.button("Download revision IMX", icon="download", on_click=self._on_download_result)\
                .props("outline")

            with ui.row():
                self.diff_button = ui.button(
                    "Create And Download Diff Report", icon="download", on_click=self._on_generate_diff
                ).props("outline")
                self.diff_spinner = ui.spinner(size="lg")
                self.diff_spinner.visible = False

            with ui.stepper_navigation():
                ui.button("Finish", on_click=self._on_finish).props("flat")
                ui.button("Back", on_click=self.stepper.previous).props("flat")

    def _build_upload_widget(self, label, on_upload, allowed_extensions: list[str] | None = None):
        accept = f"accept={','.join(allowed_extensions)}" if allowed_extensions else ""
        return ui.upload(label=label, auto_upload=True, on_upload=on_upload, max_files=1)\
            .classes("w-full").style("flex: 1").props(accept)

    def _on_imx_upload(self, event):
        self._notify(f"Uploaded IMX: {event.name}")
        self.analyze_measures_button.enable()
        self.state.imx_file_path = spooled_file_to_temp_file(event)

    def _on_json_upload(self, event):
        self._notify(f"Uploaded JSON: {event.name}")
        self.state.gr_json_file_path = spooled_file_to_temp_file(event)
        self.exclude_context_checkbox.enable()

    def _on_revision_upload(self, event):
        self.state.revisions_excel_upload_widget = spooled_file_to_temp_file(event)
        self._notify(f"Uploaded revisions Excel: {event.name}")

    def _on_download_revisions(self):
        base_name = self.state.imx_file_path.stem
        ui.download(self.state.measure_excel_file, filename=f"{base_name}-{self.state.time_stamp}-revision.xlsx")
        self._notify("Download started (revisions)")

    def _on_download_result(self):
        base_name = self.state.imx_file_path.stem
        ui.download(self.state.revision_log_zip, filename=f"{base_name}-{self.state.time_stamp}-processed.zip")
        self._notify("Download started (revision log)")

    async def _on_generate_diff(self):
        self.diff_button.disable()
        self.diff_spinner.visible = True
        try:
            temp_dir = tempfile.mkdtemp()
            self._temp_dirs.append(temp_dir)
            output_path = Path(temp_dir) / "output"
            output_path.mkdir(parents=True, exist_ok=True)

            await asyncio.to_thread(
                write_diff_output_files,
                self.state.loaded_imx_data.path,
                self.state.processed_imx,
                output_path,
                self.DEFAULT_SITUATION,
                self.DEFAULT_SITUATION,
                False,
                False,
            )

            base_name = self.state.imx_file_path.stem
            zip_name = f"{base_name}-{self.state.time_stamp}-diff.zip"
            zip_path = Path(tempfile.gettempdir()) / zip_name

            with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zipf:
                for file in output_path.rglob("*"):
                    zipf.write(file, file.relative_to(output_path))

            ui.download(zip_path, filename=zip_name)
            self._notify("Diff report ready!", type_="positive")
        except Exception as e:
            self._notify(f"Error generating diff: {e}", type_="negative")
        finally:
            self.diff_spinner.visible = False
            self.diff_button.enable()

    async def run_measure_check(self):
        self.state.time_stamp = create_timestamp()
        self.upload_spinner.visible = True
        self.analyze_measures_button.disable()
        try:
            threshold = self.threshold_input_field.value
            self.state.loaded_imx_data = await asyncio.to_thread(
                load_imxinsights_container_or_file,
                self.state.imx_file_path,
                self.DEFAULT_SITUATION,
            )
            self.state.measure_excel_file = Path(tempfile.gettempdir()) / f"measure_output_{self.state.imx_file_path.stem}.xlsx"
            if self.state.measure_excel_file.exists():
                self.state.measure_excel_file.unlink()

            await asyncio.to_thread(
                generate_measure_excel,
                self.state.loaded_imx_data,
                self.state.measure_excel_file,
                threshold if threshold else None,
            )

            if self.state.gr_json_file_path:
                puics_false, puics_true = self._classify_puics(self.state.gr_json_file_path)
                self._update_excel_puics(self.state.measure_excel_file, puics_false, puics_true)

            self.stepper.next()
        except Exception as e:
            self._notify(f"Error during measure check: {e}", type_="negative")
        finally:
            self.upload_spinner.visible = False
            self.analyze_measures_button.enable()

    async def process_revisions(self):
        self.review_spinner.visible = True
        self.process_revisions_button.disable()
        self.process_revisions_back.disable()
        try:
            tmpdir = tempfile.mkdtemp()
            self._temp_dirs.append(tmpdir)
            out_path = Path(tmpdir)

            settings = self.metadata_ui.get_metadata_settings()

            await asyncio.to_thread(
                process_imx_revisions,
                self.state.loaded_imx_data.path,
                self.state.revisions_excel_upload_widget,
                out_path,
                settings.set_metadata,
                settings.add_metadata,
                settings.source,
                settings.origin,
                settings.set_parents,
                settings.registration_time,
                True,
            )

            path_to_get = self.state.imx_file_path
            zip_path = out_path.with_suffix(".zip")
            shutil.make_archive(zip_path.with_suffix(""), "zip", out_path)

            with zipfile.ZipFile(zip_path) as zf:
                p = Path(path_to_get.name)
                internal_name = p.with_name(f"{p.stem}-processed{p.suffix}")

                if internal_name.name in zf.namelist():
                    temp_dir = tempfile.mkdtemp()
                    self._temp_dirs.append(temp_dir)
                    zf.extract(internal_name.name, path=temp_dir)
                    self.state.processed_imx = Path(temp_dir) / internal_name
                else:
                    raise FileNotFoundError(f"{internal_name} not found in {zip_path}")

            self._notify("Revisions applied successfully!", type_="positive")
            self.state.revision_log_zip = zip_path
            self.stepper.next()
        except Exception as e:
            self._notify(f"Error processing revisions: {e}", type_="negative")
        finally:
            self.review_spinner.visible = False
            self.process_revisions_button.enable()
            self.process_revisions_back.enable()

    def _classify_puics(self, json_path: Path) -> tuple[list, list]:
        puics_false, puics_true = [], []
        with open(json_path, "r", encoding="utf-8") as file:
            data = json.load(file)
            for item in data["ContextArea"]:
                puics_false.append(item["Puic"])
            if self.exclude_context_checkbox.value:
                for area in ["WorkArea", "UserArea"]:
                    for item in data[area]:
                        puics_true.append(item["Puic"])
        return puics_false, puics_true

    def _update_excel_puics(self, excel_path: Path, puics_false: list, puics_true: list):
        wb = load_workbook(excel_path)
        ws = wb["revisions"]
        header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        puic_col = header.index("object_puic") + 1
        will_be_processed_col = header.index("will_be_processed") + 1
        revision_reasoning_col = header.index("revision_reasoning") + 1

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            puic_value = row[puic_col - 1].value
            if puic_value in puics_true:
                row[will_be_processed_col - 1].value = True
                row[revision_reasoning_col - 1].value = "Not a ContextArea object"
            elif puic_value in puics_false:
                row[will_be_processed_col - 1].value = False
                row[revision_reasoning_col - 1].value = "ContextArea objects can be corrected"

        wb.save(excel_path)

    def _cleanup(self):
        for dir_path in self._temp_dirs:
            try:
                shutil.rmtree(dir_path)
            except Exception:
                pass

    def _notify(self, message: str, type_: str = "info"):
        ui.notify(message, type=type_)

    def _on_finish(self):
        self.end_and_reset_stepper()
        self._notify("Process has been reset.", type_="positive")

    def end_and_reset_stepper(self):
        self.stepper.set_value(self.UPLOAD_STEP)
        self.imx_upload_widget.reset()
        self.gr_json_upload_widget.reset()
        self.exclude_context_checkbox.value = False
        self.state = MeasureCorrectionState()
