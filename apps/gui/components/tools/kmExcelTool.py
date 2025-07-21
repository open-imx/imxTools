import asyncio
import re
import tempfile
from pathlib import Path

from nicegui import ui
from nicegui.element import Element
from openpyxl import load_workbook
from shapely.geometry import Point
from pyproj import Transformer

from apps.gui.components.widgets.uploadFile import UploadFile
from apps.gui.helpers.km_service_manager import get_km_service
from apps.gui.helpers.io import delete_later

rd2wgs = Transformer.from_crs("EPSG:28992", "EPSG:4326", always_xy=True)
wgs2rd = Transformer.from_crs("EPSG:4326", "EPSG:28992", always_xy=True)


class KmExcelTool:
    def __init__(self, container: Element):
        with container:
            ui.label("📄 Add KM to Excel").classes("font-bold text-lg")

            self.input_upload = UploadFile(
                "Upload Excel (.xlsx)", on_change=self._on_input_upload
            )

            self.simple_checkbox = ui.checkbox(
                "Use simple display value only", value=True
            ).classes("mt-2")

            with ui.row():
                self.process_button = ui.button(
                    "Add KM", on_click=self.run_add_km
                ).classes("btn-primary mt-4")
                self.spinner = (
                    ui.spinner(size="lg").props("color=primary").classes("hidden mt-2")
                )

        self.input_file: Path | None = None

    def _on_input_upload(self, file_path: Path):
        self.input_file = file_path
        ui.notify(f"Uploaded file: {file_path.name}", type="positive")

    async def run_add_km(self):
        if not self.input_file:
            ui.notify("Please upload an Excel file first.", type="warning")
            return

        self.process_button.disable()
        self.spinner.classes(remove="hidden")

        try:
            temp_output = (
                Path(tempfile.gettempdir()) / f"km_result_{self.input_file.stem}.xlsx"
            )
            if temp_output.exists():
                temp_output.unlink()

            await asyncio.to_thread(
                self._process_excel_file, self.input_file, temp_output
            )

            ui.download(temp_output, filename=temp_output.name)
            ui.notify("✅ KM values added successfully.", type="positive")

            asyncio.create_task(delete_later(temp_output))

        except Exception as e:
            ui.notify(f"❌ Error: {e}", type="negative")

        finally:
            self.process_button.enable()
            self.spinner.classes(add="hidden")

    def _process_excel_file(self, input_path: Path, output_path: Path):
        wb = load_workbook(filename=input_path)
        km_service = get_km_service()
        use_simple = self.simple_checkbox.value

        for sheet in wb.worksheets:
            gml_cols = {}
            for col in sheet.iter_cols(1, sheet.max_column):
                header = str(col[0].value)
                if header and header.endswith("Point.gml:coordinates"):
                    gml_cols[header] = col[0].column

            if not gml_cols:
                continue

            column_map = {}
            next_col = sheet.max_column + 1

            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
                for col_name, col_idx in gml_cols.items():
                    gml_value = row[col_idx - 1].value
                    if not gml_value:
                        continue

                    old_point, new_point = self._parse_gml_coordinates(gml_value)

                    km_results = []
                    if old_point:
                        old_km = km_service.get_km(old_point.x, old_point.y)
                        km_results.append(("old", old_km))
                    if new_point:
                        new_km = km_service.get_km(new_point.x, new_point.y)
                        km_results.append(("new", new_km))

                    for version, km_result in km_results:
                        if km_result and km_result.km_measures:
                            for km_measure in km_result.km_measures:
                                lint_name = getattr(
                                    km_measure.km_lint, "name", "Unknown"
                                )
                                key = (version, lint_name)

                                if key not in column_map:
                                    if use_simple:
                                        display_col = next_col
                                        sheet.cell(
                                            row=1, column=display_col
                                        ).value = f"km_display_{version}_{lint_name}"
                                        column_map[key] = (display_col,)
                                        next_col += 1
                                    else:
                                        hm_col = next_col
                                        m_col = next_col + 1
                                        name_col = next_col + 2

                                        sheet.cell(
                                            row=1, column=hm_col
                                        ).value = f"km_hm_{version}_{lint_name}"
                                        sheet.cell(
                                            row=1, column=m_col
                                        ).value = f"km_m_{version}_{lint_name}"
                                        sheet.cell(
                                            row=1, column=name_col
                                        ).value = f"km_name_{version}_{lint_name}"

                                        column_map[key] = (hm_col, m_col, name_col)
                                        next_col += 3

                                cols = column_map[key]

                                if use_simple:
                                    display_col = cols[0]
                                    sheet.cell(
                                        row=row[0].row, column=display_col
                                    ).value = km_measure.display
                                else:
                                    hm_col, m_col, name_col = cols
                                    sheet.cell(
                                        row=row[0].row, column=hm_col
                                    ).value = getattr(km_measure, "hm", "")
                                    sheet.cell(
                                        row=row[0].row, column=m_col
                                    ).value = getattr(km_measure, "distance", "")
                                    sheet.cell(
                                        row=row[0].row, column=name_col
                                    ).value = lint_name

                        else:
                            # fallback if no KM
                            key = (version, "No_KM")
                            if key not in column_map:
                                if use_simple:
                                    display_col = next_col
                                    sheet.cell(
                                        row=1, column=display_col
                                    ).value = f"km_display_{version}_NoKM"
                                    column_map[key] = (display_col,)
                                    next_col += 1
                                else:
                                    hm_col = next_col
                                    m_col = next_col + 1
                                    name_col = next_col + 2

                                    sheet.cell(
                                        row=1, column=hm_col
                                    ).value = f"km_hm_{version}_NoKM"
                                    sheet.cell(
                                        row=1, column=m_col
                                    ).value = f"km_m_{version}_NoKM"
                                    sheet.cell(
                                        row=1, column=name_col
                                    ).value = f"km_name_{version}_NoKM"

                                    column_map[key] = (hm_col, m_col, name_col)
                                    next_col += 3

                            cols = column_map[key]

                            if use_simple:
                                display_col = cols[0]
                                sheet.cell(
                                    row=row[0].row, column=display_col
                                ).value = "No KM found"
                            else:
                                hm_col, m_col, name_col = cols
                                sheet.cell(row=row[0].row, column=hm_col).value = ""
                                sheet.cell(row=row[0].row, column=m_col).value = ""
                                sheet.cell(
                                    row=row[0].row, column=name_col
                                ).value = "No KM found"

        wb.save(output_path)

    def _parse_gml_coordinates(self, gml_string: str):
        s = gml_string.strip()

        if "->" in s:
            parts = re.split(r"\s*->\s*", s)
            old_point = self._parse_single_coords(parts[0])
            new_point = self._parse_single_coords(parts[1])
            return old_point, new_point

        if s.startswith("++"):
            return None, self._parse_single_coords(s[2:])

        if s.startswith("--"):
            return self._parse_single_coords(s[2:]), None

        p = self._parse_single_coords(s)
        return p, p

    def _parse_single_coords(self, coords_str: str):
        coords = coords_str.strip().split(",")
        if len(coords) < 2:
            return None

        try:
            rd_x = float(coords[0])
            rd_y = float(coords[1])
            return Point(rd_x, rd_y)
        except ValueError:
            return None
