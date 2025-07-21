from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from src.imxTools.revision.revision_enums import (
    RevisionColumns,
    RevisionOperationValues,
)


def get_revision_template(out_file_path: str | Path):
    out_file_path = Path(out_file_path)

    # Header row (keys) and column descriptions (values)
    columns = RevisionColumns.headers()
    first_row = [RevisionColumns.to_dict()[col] for col in columns]

    example_row_1 = {
        RevisionColumns.object_path.name: "dumy.obkect.path",
        RevisionColumns.object_puic.name: "a8cfb00e-bbb3-4a83-9783-4f94e013fa9d",
        RevisionColumns.issue_comment.name: "Unknown upgrade",
        RevisionColumns.issue_cause.name: "just a example",
        RevisionColumns.attribute_or_element.name: "RailConnectionInfo.@atMeasure",
        RevisionColumns.operation.name: "UpdateAttribute",
        RevisionColumns.value_old.name: "0",
        RevisionColumns.value_new.name: "0.255",
        RevisionColumns.will_be_processed.name: "False",
        RevisionColumns.revision_reasoning.name: "Will NOT revision ProcessingStatus = False",
    }

    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    if not ws:
        raise ValueError("No Active worksheet")

    ws.title = "revisions"

    # Write enum member names as headers
    for col_idx, col_name in enumerate(columns, start=1):
        ws.cell(row=1, column=col_idx, value=col_name)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}1"

    # Write descriptions
    for col_idx, desc in enumerate(first_row, start=1):
        ws.cell(row=2, column=col_idx, value=desc)

    # Write example rows
    for row_idx, example_row in enumerate([example_row_1], start=3):
        for col_idx, col in enumerate(columns, start=1):
            ws.cell(row=row_idx, column=col_idx, value=example_row.get(col, ""))

    # Autosize columns
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_length = max(
            (
                len(str(cell))
                for row in ws.iter_rows(
                    min_col=col_idx, max_col=col_idx, values_only=True
                )
                for cell in row
            ),
            default=0,
        )
        ws.column_dimensions[col_letter].width = max_length + 2

    # Add dropdown for operation type
    operation_dv = DataValidation(
        type="list",
        formula1=f'"{",".join(RevisionOperationValues.headers())}"',
        allow_blank=True,
        showDropDown=False,
    )
    ws.add_data_validation(operation_dv)
    operation_col_letter = get_column_letter(
        columns.index(RevisionColumns.operation.name) + 1
    )
    operation_dv.add(f"{operation_col_letter}2:{operation_col_letter}1048576")

    # Add dropdown for processing status
    status_dv = DataValidation(
        type="list",
        formula1='"True,False"',
        allow_blank=True,
        showDropDown=False,
    )
    ws.add_data_validation(status_dv)
    status_col_letter = get_column_letter(
        columns.index(RevisionColumns.will_be_processed.name) + 1
    )
    status_dv.add(f"{status_col_letter}2:{status_col_letter}1048576")

    # Save Excel file
    wb.save(out_file_path)
