import re
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from shapely import Point


class KmExcelProcessor:
    """
    Processes an Excel workbook to enrich rows with kilometer (km) measure information
    based on coordinate data in columns ending with 'Point.gml:coordinates'.

    The processor detects such columns in each worksheet and extracts RD (Rijksdriehoek) coordinates
    from their cell values. It then queries a kilometer service (`km_service`) to retrieve corresponding
    km measures. These km measures are added as new columns to the Excel sheet, either in a simplified
    format (one column per km value) or in a detailed format (separate columns for hectometer,
    meters, and km lint name), based on the `use_simple` flag.

    Coordinate formats supported:
    - Single RD coordinate: `"x,y"`
    - Transition format: `"x1,y1 -> x2,y2"`
    - Added/removed point formats: `"++x,y"` or `"--x,y"`

    Args:
        km_service: An object with a `get_km(x, y)` method that returns km measure data for given coordinates.
        use_simple: If True, writes one column per km value using its display string.
                    If False, writes three columns per km value (hm, meters, name).
    """

    def __init__(self, km_service, use_simple: bool = True):
        self.km_service = km_service
        self.use_simple = use_simple

    def process(self, input_path: Path, output_path: Path):
        wb = load_workbook(filename=input_path)

        for sheet in wb.worksheets:
            gml_columns = self._detect_gml_columns(sheet)
            if not gml_columns:
                continue

            column_map = {}
            next_col = sheet.max_column + 1

            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
                km_data = self._collect_km_measures_for_row(row, gml_columns)

                next_col = self._write_km_measures_to_row(
                    sheet, row[0].row, km_data, column_map, next_col
                )

        self._autosize_columns(wb)
        wb.save(output_path)

    def _detect_gml_columns(self, sheet):
        gml_cols = {}
        for col in sheet.iter_cols(1, sheet.max_column):
            header = str(col[0].value)
            if header and header.endswith("Point.gml:coordinates"):
                gml_cols[header] = col[0].column
        return gml_cols

    def _collect_km_measures_for_row(self, row, gml_columns):
        results = {"old": [], "new": []}

        for _, col_idx in gml_columns.items():
            gml_value = row[col_idx - 1].value
            if not gml_value:
                continue

            old_point, new_point = self._parse_gml_coordinates(gml_value)

            for version, point in [("old", old_point), ("new", new_point)]:
                if not point:
                    continue
                try:
                    km_result = self.km_service.get_km(point.x, point.y)
                    if km_result and km_result.km_measures:
                        results[version].extend(km_result.km_measures)
                    else:
                        results[version].append(None)
                except Exception as e:
                    results[version].append(f"ERROR: {e}")

        return results

    def _write_km_measures_to_row(self, sheet, row_idx, km_data, column_map, start_col):
        max_lints = max(len(km_data["old"]), len(km_data["new"]))

        for lint_index in range(max_lints):
            for version in ["old", "new"]:
                km_measure = (
                    km_data[version][lint_index]
                    if lint_index < len(km_data[version])
                    else None
                )
                key = (lint_index, version)

                if self.use_simple:
                    if key not in column_map:
                        col = start_col
                        sheet.cell(
                            row=1, column=col
                        ).value = f"km_{lint_index + 1}_{version}"
                        column_map[key] = (col,)
                        start_col += 1
                    col = column_map[key][0]
                    value = self._format_simple(km_measure)
                    sheet.cell(row=row_idx, column=col).value = value
                else:
                    if key not in column_map:
                        col1, col2, col3 = start_col, start_col + 1, start_col + 2
                        sheet.cell(
                            row=1, column=col1
                        ).value = f"km_{lint_index + 1}_{version}_hm"
                        sheet.cell(
                            row=1, column=col2
                        ).value = f"km_{lint_index + 1}_{version}_m"
                        sheet.cell(
                            row=1, column=col3
                        ).value = f"km_{lint_index + 1}_{version}_name"
                        column_map[key] = (col1, col2, col3)
                        start_col += 3

                    col1, col2, col3 = column_map[key]
                    values = self._format_detailed(km_measure)
                    sheet.cell(row=row_idx, column=col1).value = values[0]
                    sheet.cell(row=row_idx, column=col2).value = values[1]
                    sheet.cell(row=row_idx, column=col3).value = values[2]

        return start_col

    def _format_simple(self, km_measure):
        if isinstance(km_measure, str):
            return km_measure
        if km_measure:
            return km_measure.display
        return "No KM found"

    def _format_detailed(self, km_measure):
        if isinstance(km_measure, str):
            return km_measure, "", ""
        if km_measure:
            return (
                getattr(km_measure, "hm", ""),
                getattr(km_measure, "distance", ""),
                getattr(km_measure.km_lint, "name", ""),
            )
        return "", "", "No KM found"

    def _autosize_columns(self, workbook):
        for sheet in workbook.worksheets:
            for column_cells in sheet.columns:
                max_length = 0
                col_letter = get_column_letter(column_cells[0].column)
                for cell in column_cells:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except Exception:
                        pass
                sheet.column_dimensions[col_letter].width = max_length + 2

    def _parse_gml_coordinates(self, gml_string: str):
        s = gml_string.strip()

        if "->" in s:
            parts = re.split(r"\s*->\s*", s)
            return self._parse_single_coords(parts[0]), self._parse_single_coords(
                parts[1]
            )

        if s.startswith("++"):
            return None, self._parse_single_coords(s[2:])
        if s.startswith("--"):
            return self._parse_single_coords(s[2:]), None

        p = self._parse_single_coords(s)
        return p, p

    def _parse_single_coords(self, coords_str: str):
        coords = coords_str.strip().split(",")
        if len(coords) < 2:
            return None
        try:
            rd_x = float(coords[0])
            rd_y = float(coords[1])
            return Point(rd_x, rd_y)
        except ValueError:
            return None
