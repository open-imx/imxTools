import re

from openpyxl import Workbook
from openpyxl.cell import Cell, MergedCell
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


def get_cell_background_color(cell: Cell | MergedCell) -> str | None:
    """Extracts the background color (last 6 hex chars) of a cell, if RGB color is set."""
    if cell.fill and cell.fill.fgColor and cell.fill.fgColor.type == "rgb":
        return cell.fill.fgColor.rgb[-6:]
    return None


def get_column_indices(
    ws: Worksheet, header_row: int, keys: list[str]
) -> dict[str, int]:
    """Maps the column names in the header row to their column index for the given keys."""
    header_values = {
        str(ws.cell(row=header_row, column=col).value): col
        for col in range(1, ws.max_column + 1)
    }
    return {key: header_values[key] for key in keys if key in header_values}


def _set_comment_fill(
    ws: Worksheet, row_idx: int, col: str | int, bg_color: str | None
) -> None:
    """Sets cell fill color in the given column and row index.

    Args:
        ws: The worksheet to modify.
        row_idx: The row index (1-based).
        col: The column (letter as string or index as int).
        bg_color: The background color to apply, if any.
    """
    if not bg_color:
        return

    if isinstance(col, int):
        col_letter = get_column_letter(col)
    else:
        col_letter = col

    cell = ws[f"{col_letter}{row_idx}"]
    cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")


def get_display_text(cell):
    """Returns the visible text in a cell, including formulas using HYPERLINK()."""
    if cell.hyperlink and cell.value is None:
        return (
            str(cell.hyperlink.display)
            if cell.hyperlink.display
            else str(cell.hyperlink.target)
        )

    if isinstance(cell.value, str) and cell.value.startswith("=HYPERLINK"):
        match = re.match(
            r'=HYPERLINK\(([^,]+),\s*("?)(.*?)\2\)', cell.value, re.IGNORECASE
        )
        if match:
            return match.group(3)
        return cell.value

    return str(cell.value) if cell.value is not None else ""


def auto_size_columns(ws):
    for col_cells in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            text = get_display_text(cell)
            if "\n" in text:
                # Consider multiline text
                max_line = max(len(line) for line in text.splitlines())
            else:
                max_line = len(text)
            max_length = max(max_length, max_line)
        ws.column_dimensions[col_letter].width = max_length + 2  # Padding


def move_sheet_after(workbook: Workbook, sheet_to_move, after_title: str) -> None:
    """
    Moves the given worksheet to a position immediately after the sheet with the specified title.
    If the target sheet is not found, inserts at the beginning.

    Args:
        workbook: The workbook containing the worksheets.
        sheet_to_move: The worksheet to move.
        after_title: The title of the sheet after which to insert the moved sheet.
    """
    worksheets = workbook.worksheets

    if sheet_to_move not in worksheets:
        return

    current_index = worksheets.index(sheet_to_move)

    try:
        after_index = next(
            i for i, ws in enumerate(worksheets) if ws.title == after_title
        )
    except StopIteration:
        after_index = -1  # Insert at beginning

    # Adjust after_index if sheet_to_move is before the target
    if current_index <= after_index:
        target_index = after_index
    else:
        target_index = after_index + 1

    offset = target_index - current_index
    if offset != 0:
        workbook.move_sheet(sheet_to_move, offset)


def add_header_and_auto_filter(
    ws: Worksheet, headers: list[str], row: int = 1, freeze: bool = True
) -> None:
    """
    Inserts headers at the specified row, applies an autofilter to that row,
    and optionally freezes panes just below the header.

    Args:
        ws: The worksheet to modify.
        headers: A list of header column names.
        row: The row number at which to insert the headers and apply autofilter.
        freeze: Whether to freeze panes just below the header row.
    """
    if not headers:
        return

    for col_index, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_index)
        cell.value = header

    end_col = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A{row}:{end_col}{row}"

    if freeze:
        ws.freeze_panes = f"A{row + 1}"
