import os
import shutil
from pathlib import Path

from imxInsights.utils.report_helpers import REVIEW_STYLES, add_review_styles_to_excel
from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell, MergedCell
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from src.imxTools.comments.comment_entry import CommentEntry
from src.imxTools.comments.comments_enums import CommentColumns
from src.imxTools.comments.openpyxl_helpers import (
    _set_comment_fill,
    add_header_and_auto_filter,
    auto_size_columns,
    get_cell_background_color,
    get_column_indices,
    move_sheet_after,
)
from src.imxTools.settings import config


# todo: better naming: something like get diff context
def get_cell_context(
    ws: Worksheet, row_idx: int, columns: dict[str, int]
) -> dict[str, str]:
    """Extracts contextual information for a row based on column indices."""

    def get(col_key: str) -> str:
        col = columns.get(col_key)
        if col is None:
            return ""
        val = ws.cell(row=row_idx, column=col).value
        return str(val) if val is not None else ""

    return {
        CommentColumns.object_puic: get("@puic"),
        CommentColumns.object_path: get("path"),
        CommentColumns.change_status: get("status"),
        CommentColumns.geometry_status: get("geometry_status"),
    }


def build_comment_entry(
    cell: Cell | MergedCell,
    header: str,
    sheet_name: str,
    comment_text: str,
    comment_row: int | None,
    context: dict[str, str],
) -> CommentEntry:
    """Creates a CommentEntry object from a cell and its context."""
    return CommentEntry(
        sheet_name=sheet_name,
        header_value=header,
        cell_value=str(cell.value) if cell.value is not None else "",
        comment=comment_text,
        cell_address=cell.coordinate,
        bg_color=get_cell_background_color(cell),
        row=comment_row or cell.row or 0,
        column=cell.column or 0,
        puic=context[CommentColumns.object_puic],
        path=context[CommentColumns.object_path],
        status=context[CommentColumns.change_status],
        geometry_status=context[CommentColumns.geometry_status],
    )


def handle_header_comment(
    cell: Cell | MergedCell,
    ws: Worksheet,
    header_value: str,
    columns: dict[str, int],
    sheet_name: str,
    header_row: int,
) -> tuple[list[CommentEntry], list[str]]:
    """
    Handles comments placed in a header cell.
    These comments are inherited by cells underneath that share the same color.
    # TODO: if header, its for all except those whit a comment it self.
    """
    if get_cell_background_color(cell) == "FFFFFF":
        return [], []  # Skip plain white background cells

    comments = []
    inherited_cells = []

    for row_idx in range(header_row + 1, ws.max_row + 1):
        if cell.column is None:
            continue
        target_cell = ws.cell(row=row_idx, column=cell.column)
        if target_cell.comment or not target_cell.value:
            continue
        color = get_cell_background_color(target_cell)
        if color and color != "000000":
            context = get_cell_context(ws, row_idx, columns)
            if cell.comment:
                entry = build_comment_entry(
                    target_cell,
                    header_value,
                    sheet_name,
                    cell.comment.text,
                    header_row,
                    context,
                )
                comments.append(entry)
                inherited_cells.append(target_cell.coordinate)

    return comments, inherited_cells


def handle_data_comment(
    cell: Cell | MergedCell,
    ws: Worksheet,
    header_value: str,
    columns: dict[str, int],
    sheet_name: str,
) -> list[CommentEntry]:
    """Processes a direct comment found in a cell (non-header)."""
    if cell.row:
        context = get_cell_context(ws, cell.row, columns)
        return [
            build_comment_entry(
                cell,
                header_value,
                sheet_name,
                cell.comment.text if cell.comment else "",
                None,
                context,
            )
        ]
    return []


def _build_comment_row(entry: CommentEntry, color_to_status: dict[str, str]) -> list:
    """Builds a row for writing to the comments sheet."""
    link_text = color_to_status.get(str(entry.bg_color), "link")
    link = f'=HYPERLINK("#\'{entry.sheet_name}\'!{entry.cell_address}", "{link_text}")'
    return [
        entry.path,
        entry.puic,
        entry.status,
        entry.geometry_status,
        link,
        entry.header_value,
        entry.cell_value,
        entry.comment,
        entry.bg_color,
        entry.sheet_name,
        entry.row,
        entry.column,
    ]


def write_comments_sheet(
    wb: Workbook, comments: list[CommentEntry], overwrite: bool = False
) -> None:
    """
    Creates or overwrites the 'comments' sheet and populates it with extracted comment entries.
    Applies background color and adds hyperlinks to original cells.
    """
    if config.ISSUE_LIST_SHEET_NAME in wb.sheetnames:
        if overwrite:
            del wb[config.ISSUE_LIST_SHEET_NAME]
        else:
            raise ValueError(
                f"Sheet '{config.ISSUE_LIST_SHEET_NAME}' already exists. Set overwrite=True to overwrite."
            )

    ws_comments = wb.create_sheet(config.ISSUE_LIST_SHEET_NAME)

    move_sheet_after(wb, ws_comments, after_title="info")
    add_header_and_auto_filter(ws_comments, CommentColumns.names(), row=1)

    # Add each comment as a row with formatting
    color_to_status = {v: k for k, v in REVIEW_STYLES.items()}
    for entry in comments:
        row = _build_comment_row(entry, color_to_status)
        ws_comments.append(row)
        _set_comment_fill(ws_comments, ws_comments.max_row, "E", entry.bg_color)

    auto_size_columns(ws_comments)
    comment_idx = wb.sheetnames.index(config.ISSUE_LIST_SHEET_NAME)
    wb.active = comment_idx  # type: ignore
    for i, ws in enumerate(wb.worksheets):
        ws.sheet_view.tabSelected = i == comment_idx


def extract_comments_to_new_sheet(
    file_path: str | Path,
    output_path: str | None = None,
    header_row: int = 1,
    add_to_wb: bool = False,
    overwrite: bool = False,
) -> None:
    """
    Extracts all comments (direct and inherited) from a workbook and writes them to a new 'comments' sheet.
    Depending on the flags, either adds to the same workbook or a new one.
    """
    if not output_path and not add_to_wb:
        raise ValueError("When adding to an existing workbook, provide an output path.")

    # Decide whether to write to a new file or modify in-place
    target_path = output_path if output_path and add_to_wb else file_path
    if output_path and add_to_wb:
        if os.path.exists(output_path) and not overwrite:
            raise FileExistsError(
                f"File '{output_path}' already exists. Set overwrite=True."
            )
        shutil.copyfile(file_path, output_path)

    wb = load_workbook(target_path, data_only=True)
    all_comments = []
    inherited_cells = set()

    # Process each worksheet
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        columns = get_column_indices(
            ws, header_row, ["@puic", "path", "status", "geometry_status"]
        )

        for row in ws.iter_rows():
            for cell in row:
                if cell.column is None or cell.row is None:
                    continue

                col_letter = get_column_letter(cell.column)
                header_cell = ws[f"{col_letter}{header_row}"]
                header_value = str(header_cell.value or "")

                if cell.comment:
                    if cell.row == header_row:  # Process header-level comments
                        comments, inherited = handle_header_comment(
                            cell, ws, header_value, columns, sheet_name, header_row
                        )
                        all_comments.extend(comments)
                        inherited_cells.update(inherited)
                    else:
                        all_comments.extend(
                            handle_data_comment(
                                cell, ws, header_value, columns, sheet_name
                            )
                        )
                else:
                    # Process inherited comments (colored cells without comment)
                    color = get_cell_background_color(cell)
                    if (
                        cell.row > header_row
                        and color in REVIEW_STYLES.values()
                        and cell.value
                        and cell.coordinate not in inherited_cells
                    ):
                        context = get_cell_context(ws, cell.row, columns)
                        entry = build_comment_entry(
                            cell, header_value, sheet_name, "", None, context
                        )
                        all_comments.append(entry)

    # Write comments to workbook
    if add_to_wb:
        if os.path.exists(file_path) and not overwrite:
            raise FileExistsError(
                f"File '{file_path}' already exists. Set overwrite=True."
            )
        write_comments_sheet(wb, all_comments, overwrite)
        wb.save(target_path)

        add_review_styles_to_excel(target_path)

    else:
        output_path = output_path or str(file_path).replace(".xlsx", "_comments.xlsx")
        wb_new = Workbook()
        if wb_new.active:
            wb_new.remove(wb_new.active)
        write_comments_sheet(wb_new, all_comments, overwrite)
        wb_new.save(output_path)

        add_review_styles_to_excel(output_path)
