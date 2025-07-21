from copy import copy
from pathlib import Path
from typing import Any, cast

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from src.imxTools.comments.comments_enums import CommentColumns
from src.imxTools.settings import config
from src.imxTools.utils.helpers import ensure_paths


def copy_full_sheet(source_ws: Worksheet, target_ws: Worksheet) -> None:
    for row in source_ws.iter_rows():
        for cell in row:
            if cell.row is None or cell.column is None:
                continue

            target_cell = target_ws.cell(row=cell.row, column=cell.column)

            if cell.data_type == "f":
                target_cell.value = f"={cell.value}"
            else:
                target_cell.value = cell.value

            if cell.hyperlink:
                target_cell.hyperlink = copy(cell.hyperlink)

            if cell.has_style:
                target_cell.font = cast(Any, copy(cell.font))
                target_cell.fill = cast(Any, copy(cell.fill))
                target_cell.border = cast(Any, copy(cell.border))
                target_cell.alignment = cast(Any, copy(cell.alignment))
                target_cell.number_format = cell.number_format
                target_cell.protection = cast(Any, copy(cell.protection))

    for col_letter, col_dim in source_ws.column_dimensions.items():
        if col_dim.width is not None:
            target_ws.column_dimensions[col_letter].width = col_dim.width

    for row_idx, row_dim in source_ws.row_dimensions.items():
        if row_dim.height is not None:
            target_ws.row_dimensions[row_idx].height = row_dim.height

    for merged_range in source_ws.merged_cells.ranges:
        target_ws.merge_cells(str(merged_range))


def get_sheet_headers(ws: Worksheet, header_row: int = 1) -> dict[str, int]:
    return {
        str(cell.value): idx
        for idx, cell in enumerate(
            next(ws.iter_rows(min_row=header_row, max_row=header_row)), start=1
        )
        if cell.value is not None
    }


def find_column_by_value(ws: Worksheet, target: str, header_row: int) -> int | None:
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=header_row, column=col).value == target:
            return col
    return None


def find_row_by_value(
    ws: Worksheet, col: int, value: Any, start_row: int
) -> int | None:
    for row in range(start_row, ws.max_row + 1):
        if str(ws.cell(row=row, column=col).value) == str(value):
            return row
    return None


def create_summary_sheet(
    wb: Workbook,
    processed: list[dict[str, Any]],
    skipped: list[dict[str, Any]],
    not_found: list[dict[str, Any]],
) -> Worksheet:
    summary_ws = wb.create_sheet("CommentPlacementSummary")
    summary_ws.append(
        ["Status", "Sheet", "ImxPath", "Puic", "Value", "Comment", "Reason"]
    )

    for status, rows in [
        ("Placed", processed),
        ("Skipped", skipped),
        ("Failed", not_found),
    ]:
        for row in rows:
            summary_ws.append(
                [
                    status,
                    row.get("CommentSheetName", ""),
                    row.get("ImxPath", ""),
                    row.get("Puic", ""),
                    row.get("Value", ""),
                    row.get("Comment", ""),
                    row.get("Reason", ""),
                ]
            )
    return summary_ws


def extract_display_text(formula: str) -> str:
    parts = formula.split(",")
    if len(parts) >= 2:
        return parts[1].strip().rstrip(')"').strip('"')
    return ""


def apply_comment_to_cell(
    ws: Worksheet,
    header_col: int,
    puic_col: int,
    header_row: int,
    puic: Any,
    comment_text: str,
    data: dict[str, Any],
    processed: list[dict[str, Any]],
    skipped: list[dict[str, Any]],
    not_found: list[dict[str, Any]],
) -> None:
    target_row = find_row_by_value(ws, puic_col, puic, header_row + 1)
    if target_row is None:
        not_found.append({**data, "Reason": f"Puic '{puic}' not found"})
        return

    cell = ws.cell(row=target_row, column=header_col)

    style_name = extract_display_text(
        str(data.get(CommentColumns.comment_type.name, ""))
    )
    parent_wb = ws.parent if isinstance(ws.parent, Workbook) else None
    if parent_wb and style_name in parent_wb.named_styles:
        cell.style = style_name
    else:
        print(f"Style '{style_name}' not found.")

    is_header_comment = data.get(CommentColumns.comment_row) == header_row
    if is_header_comment:
        header_cell = ws.cell(row=header_row, column=header_col)
        existing_comment = header_cell.comment
        existing_comment_text = (
            existing_comment.text.strip() if existing_comment else ""
        )

        if comment_text and comment_text != existing_comment_text:
            new_comment = Comment(comment_text, "IMX Tool")
            new_comment.visible = False
            header_cell.comment = new_comment

        header_cell.style = style_name

    header_comment = ws.cell(row=header_row, column=header_col).comment
    header_comment_text = header_comment.text.strip() if header_comment else ""

    if not is_header_comment:
        if comment_text == header_comment_text:
            processed.append(
                {
                    **data,
                    "Reason": "header comment, only cell color",
                    "Value": cell.value,
                }
            )
        elif comment_text:
            new_comment = Comment(comment_text, "open-imx-comment-replacer")
            new_comment.visible = False
            cell.comment = new_comment
            processed.append({**data, "Value": cell.value})
        else:
            skipped.append({**data, "Value": cell.value, "Reason": "Empty comment"})
    else:
        if comment_text:
            processed.append({**data, "Value": cell.value})
        else:
            skipped.append(
                {**data, "Value": cell.value, "Reason": "Empty header comment"}
            )


def auto_resize_columns(ws: Worksheet) -> None:
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column if col and col[0].column else 1)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2


def apply_comments_from_issue_list(
    issue_list_path: str | Path,
    new_diff_path: str | Path,
    output_path: str | Path,
    header_row: int = 1,
) -> None:
    issue_list_path, new_diff_path, output_path = ensure_paths(
        issue_list_path, new_diff_path, output_path
    )

    issue_wb = load_workbook(issue_list_path, data_only=False)
    diff_wb = load_workbook(new_diff_path)

    if config.ISSUE_LIST_SHEET_NAME not in issue_wb.sheetnames:
        raise ValueError(
            f"No '{config.ISSUE_LIST_SHEET_NAME}' sheet found in the issue list workbook."
        )

    issue_ws = issue_wb[config.ISSUE_LIST_SHEET_NAME]
    headers = get_sheet_headers(issue_ws, header_row)

    processed: list[dict[str, Any]] = []
    skipped: list[dict[str, Any]] = []
    not_found: list[dict[str, Any]] = []

    all_rows: list[dict[str, Any]] = []
    for row in issue_ws.iter_rows(min_row=header_row + 1, values_only=True):
        data = {
            str(key): row[idx - 1] if idx - 1 < len(row) else None
            for key, idx in headers.items()
        }
        all_rows.append(data)

    def safe_int(val: Any) -> int:
        try:
            return int(val)
        except Exception:
            return 1_000_000_000

    all_rows.sort(key=lambda d: safe_int(d.get(CommentColumns.comment_row)))

    for data in all_rows:
        try:
            sheetname = str(data.get(CommentColumns.comment_sheet_name.name))
            imx_path = str(data.get(CommentColumns.header_value.name))
            puic = data.get(CommentColumns.object_puic.name)
            comment_val = data.get(CommentColumns.comment.name)
            comment_text = str(comment_val).strip() if comment_val else ""

            if not all([sheetname, imx_path, puic]):
                not_found.append(
                    {**data, "Reason": "Missing CommentSheetName, ImxPath, or Puic"}
                )
                continue

            if sheetname not in diff_wb.sheetnames:
                not_found.append(
                    {**data, "Reason": f"Sheet '{sheetname}' not found in new diff"}
                )
                continue

            ws = diff_wb[sheetname]

            header_col = find_column_by_value(ws, imx_path, header_row)
            if header_col is None:
                not_found.append(
                    {**data, "Reason": f"ImxPath '{imx_path}' not found in header"}
                )
                continue

            puic_col = find_column_by_value(ws, "@puic", header_row)
            if puic_col is None:
                not_found.append({**data, "Reason": "@puic column not found"})
                continue

            apply_comment_to_cell(
                ws,
                header_col,
                puic_col,
                header_row,
                puic,
                comment_text,
                data,
                processed,
                skipped,
                not_found,
            )

        except Exception as e:
            not_found.append({**data, "Reason": f"Unexpected error: {str(e)}"})

    issue_list_ws = diff_wb.create_sheet(config.ISSUE_LIST_SHEET_NAME)
    copy_full_sheet(issue_ws, issue_list_ws)

    ordered_titles = ["info", config.ISSUE_LIST_SHEET_NAME, "CommentPlacementSummary"]
    ordered_sheets = []
    for title in ordered_titles:
        if title in diff_wb.sheetnames:
            ordered_sheets.append(diff_wb[title])  # type: ignore[arg-type]
    for sheet in diff_wb.worksheets:
        if sheet not in ordered_sheets:
            ordered_sheets.append(sheet)  # type: ignore[arg-type]
    diff_wb._sheets = ordered_sheets  # type: ignore[attr-defined]

    diff_wb.active = diff_wb.sheetnames.index("info")  # type: ignore[assignment]
    for sheet in diff_wb.worksheets:
        sheet.sheet_view.tabSelected = sheet.title == "info"

    create_summary_sheet(diff_wb, processed, skipped, not_found)

    diff_wb.save(output_path)
    print(f"✅ Comments copied and saved to '{output_path}'")
    print(
        f"Summary: {len(processed)} placed, {len(skipped)} skipped, {len(not_found)} failed."
    )
